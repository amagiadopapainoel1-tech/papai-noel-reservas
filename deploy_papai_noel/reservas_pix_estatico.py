from flask import Flask, render_template
from openpyxl import load_workbook

app = Flask(__name__)

# Caminho da planilha
ARQUIVO_EXCEL = "Agenda_Papai_Noel_Atualizada.xlsx"

@app.get("/")
def index():
    return "<h1>Sistema do Papai Noel 🎅<br>Use /agenda para ver os atendimentos.</h1>"

@app.get("/agenda")
def agenda():
    # Carregar a planilha
    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb.active

    dados = []

    # Ignorar cabeçalho (linha 1)
    for linha in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in linha):
            continue  # Ignora linha vazia
        dados.append(linha)

    return render_template("agenda.html", dados=dados)

# Suas rotas antigas continuam funcionando
@app.get('/run_process')
def run_process():
    return "Processamento iniciado com sucesso!"

@app.get('/confirmar_pagamento')
def confirmar_pagamento():
    return "Pagamento confirmado!"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
